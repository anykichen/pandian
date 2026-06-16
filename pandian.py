try:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
    HAS_TK = True
except Exception:
    HAS_TK = False
import argparse
import openpyxl
import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import unicodedata

# -------------------------- 数据库初始化 --------------------------
def init_db():
    conn = sqlite3.connect('fixed_asset.db')
    c = conn.cursor()
    # 固定资产主表
    c.execute('''
        CREATE TABLE IF NOT EXISTS assets (
            main_asset_id TEXT PRIMARY KEY,
            asset_name TEXT,
            owner_name TEXT,
            location_2025 TEXT,
            location_2026 TEXT DEFAULT '',
            is_checked INTEGER DEFAULT 0
        )
    ''')
    conn.commit()
    conn.close()

# -------------------------- 列宽自动适配函数 --------------------------
def display_width(text):
    return sum(2 if unicodedata.east_asian_width(c) in ('F','W') else 1 for c in str(text or ''))

def auto_fit_columns(ws, min_w=8, max_w=50, padding=3):
    for col_cells in ws.columns:
        letter = col_cells[0].column_letter
        try:
            merged_cls = openpyxl.cell.cell.MergedCell
        except Exception:
            merged_cls = None
        w = max((display_width(c.value) for c in col_cells
                 if not (merged_cls is not None and isinstance(c, merged_cls)) and c.value is not None), default=0)
        ws.column_dimensions[letter].width = max(min_w, min(w * 1.1 + padding, max_w))

# -------------------------- 主应用类 --------------------------
class AssetCheckApp:
    def __init__(self, root):
        self.root = root
        self.root.title("固定资产盘点系统")
        self.root.geometry("600x400")
        self.root.resizable(False, False)

        # 初始化数据库
        init_db()

        # 界面布局
        self.create_widgets()

    def create_widgets(self):
        # 1. 数据导入区域
        import_frame = ttk.LabelFrame(self.root, text="1. 固定资产数据导入")
        import_frame.pack(fill="x", padx=20, pady=10)

        self.import_path = tk.StringVar()
        ttk.Entry(import_frame, textvariable=self.import_path, width=50).grid(row=0, column=0, padx=10, pady=10)
        ttk.Button(import_frame, text="选择Excel文件", command=self.select_file).grid(row=0, column=1, padx=5, pady=10)
        ttk.Button(import_frame, text="一键导入", command=self.import_data).grid(row=0, column=2, padx=5, pady=10)

        # 2. 盘点操作区域
        check_frame = ttk.LabelFrame(self.root, text="2. 资产盘点操作")
        check_frame.pack(fill="x", padx=20, pady=10)

        # 主资产编号输入
        ttk.Label(check_frame, text="主资产编号：").grid(row=0, column=0, padx=10, pady=10, sticky="e")
        self.asset_id_var = tk.StringVar()
        self.asset_id_entry = ttk.Entry(check_frame, textvariable=self.asset_id_var, width=20)
        self.asset_id_entry.grid(row=0, column=1, padx=5, pady=10)
        ttk.Button(check_frame, text="查询资产", command=self.query_asset).grid(row=0, column=2, padx=5, pady=10)

        # 资产信息回显
        ttk.Label(check_frame, text="资产名称：").grid(row=1, column=0, padx=10, pady=5, sticky="e")
        self.asset_name_var = tk.StringVar()
        ttk.Entry(check_frame, textvariable=self.asset_name_var, state="readonly", width=30).grid(row=1, column=1, padx=5, pady=5, columnspan=2)

        ttk.Label(check_frame, text="姓名：").grid(row=2, column=0, padx=10, pady=5, sticky="e")
        self.owner_name_var = tk.StringVar()
        ttk.Entry(check_frame, textvariable=self.owner_name_var, state="readonly", width=30).grid(row=2, column=1, padx=5, pady=5, columnspan=2)

        ttk.Label(check_frame, text="2025位置：").grid(row=3, column=0, padx=10, pady=5, sticky="e")
        self.location_2025_var = tk.StringVar()
        ttk.Entry(check_frame, textvariable=self.location_2025_var, state="readonly", width=30).grid(row=3, column=1, padx=5, pady=5, columnspan=2)

        # 2026位置输入
        ttk.Label(check_frame, text="2026位置：").grid(row=4, column=0, padx=10, pady=10, sticky="e")
        self.location_2026_var = tk.StringVar()
        ttk.Entry(check_frame, textvariable=self.location_2026_var, width=30).grid(row=4, column=1, padx=5, pady=10, columnspan=2)
        ttk.Button(check_frame, text="保存盘点结果", command=self.save_check_result).grid(row=4, column=3, padx=5, pady=10)

        # 3. 报表生成区域
        report_frame = ttk.LabelFrame(self.root, text="3. 盘点报表生成")
        report_frame.pack(fill="x", padx=20, pady=10)

        ttk.Button(report_frame, text="导出已盘点资产报表", command=self.export_checked).pack(side="left", padx=20, pady=10)
        ttk.Button(report_frame, text="导出未盘点资产报表", command=self.export_unchecked).pack(side="left", padx=20, pady=10)
        ttk.Button(report_frame, text="导出全量盘点报表", command=self.export_all).pack(side="left", padx=20, pady=10)

    # -------------------------- 功能函数 --------------------------
    def select_file(self):
        file_path = filedialog.askopenfilename(
            title="选择固定资产Excel文件",
            filetypes=[("Excel文件", "*.xlsx;*.xls")]
        )
        if file_path:
            self.import_path.set(file_path)

    def import_data(self):
        file_path = self.import_path.get()
        if not file_path:
            messagebox.showwarning("警告", "请先选择Excel文件")
            return

        try:
            # 读取Excel文件，匹配您的模板列名
            df = pd.read_excel(file_path, dtype={'主资产编号': str})
            required_cols = ['主资产编号', '资产名称', '姓名', '2025位置']
            if not all(col in df.columns for col in required_cols):
                messagebox.showerror("错误", "Excel文件列名不符合模板要求！\n必须包含：主资产编号、资产名称、姓名、2025位置")
                return

            # 处理2026位置列，没有则新增
            if '2026位置' not in df.columns:
                df['2026位置'] = ''

            # 去重，保留主资产编号唯一
            df = df.drop_duplicates(subset=['主资产编号'], keep='first')

            # 写入数据库
            conn = sqlite3.connect('fixed_asset.db')
            df.to_sql('assets', conn, if_exists='replace', index=False)
            # 补充盘点状态字段
            c = conn.cursor()
            c.execute('UPDATE assets SET is_checked = 1 WHERE location_2026 IS NOT NULL AND location_2026 != ""')
            conn.commit()
            conn.close()

            messagebox.showinfo("成功", f"数据导入完成！共导入 {len(df)} 条固定资产数据")
            self.import_path.set("")

        except Exception as e:
            messagebox.showerror("错误", f"导入失败：{str(e)}")

    def query_asset(self):
        asset_id = self.asset_id_var.get().strip()
        if not asset_id:
            messagebox.showwarning("警告", "请输入主资产编号")
            return

        # 清空之前的内容
        self.asset_name_var.set("")
        self.owner_name_var.set("")
        self.location_2025_var.set("")
        self.location_2026_var.set("")

        # 查询数据库
        conn = sqlite3.connect('fixed_asset.db')
        c = conn.cursor()
        c.execute('SELECT asset_name, owner_name, location_2025, location_2026 FROM assets WHERE main_asset_id = ?', (asset_id,))
        result = c.fetchone()
        conn.close()

        if result:
            self.asset_name_var.set(result[0])
            self.owner_name_var.set(result[1])
            self.location_2025_var.set(result[2])
            self.location_2026_var.set(result[3] if result[3] else "")
        else:
            messagebox.showinfo("提示", "未找到该主资产编号对应的资产")

    def save_check_result(self):
        asset_id = self.asset_id_var.get().strip()
        location_2026 = self.location_2026_var.get().strip()

        if not asset_id:
            messagebox.showwarning("警告", "请先查询资产信息")
            return
        if not location_2026:
            messagebox.showwarning("警告", "请输入2026位置")
            return

        # 保存到数据库
        try:
            conn = sqlite3.connect('fixed_asset.db')
            c = conn.cursor()
            c.execute('UPDATE assets SET location_2026 = ?, is_checked = 1 WHERE main_asset_id = ?', (location_2026, asset_id))
            conn.commit()
            conn.close()

            messagebox.showinfo("成功", "盘点结果已保存！")
            # 清空输入，准备下一次盘点
            self.asset_id_var.set("")
            self.asset_name_var.set("")
            self.owner_name_var.set("")
            self.location_2025_var.set("")
            self.location_2026_var.set("")
            # 光标回到主资产编号输入框
            self.asset_id_entry.focus()

        except Exception as e:
            messagebox.showerror("错误", f"保存失败：{str(e)}")

    # -------------------------- 报表导出函数 --------------------------
    def export_excel(self, data, title, file_name):
        if len(data) == 0:
            messagebox.showinfo("提示", "暂无符合条件的资产数据")
            return

        # 选择保存路径
        save_path = filedialog.asksaveasfilename(
            title=f"保存{title}",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx")],
            initialfile=file_name
        )
        if not save_path:
            return

        # 生成Excel文件，符合可视化规范
        wb = Workbook()
        ws = wb.active
        ws.title = title

        # 表头
        headers = ['主资产编号', '资产名称', '姓名', '2025位置', '2026位置', '盘点状态']
        header_fill = PatternFill('solid', fgColor='0070C0')
        header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        center_align = Alignment(horizontal='center', vertical='center')

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # 数据行
        zebra_blue = 'EBF1F8'
        white = 'FFFFFF'
        for row_idx, row_data in enumerate(data, 2):
            bg = zebra_blue if (row_idx - 2) % 2 == 0 else white
            fill = PatternFill('solid', fgColor=bg)
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='left' if col_idx in [2,3,4,5] else 'center', vertical='center')
                cell.fill = fill
                # 盘点状态标色
                if col_idx == 6:
                    if value == '已盘点':
                        cell.font = Font(bold=True, color='00B050')
                    else:
                        cell.font = Font(bold=True, color='FF0000')

        # 自动适配列宽
        auto_fit_columns(ws)
        # 冻结表头
        ws.freeze_panes = 'A2'
        # 保存文件
        wb.save(save_path)
        messagebox.showinfo("成功", f"{title}已导出到：{save_path}")

    def export_checked(self):
        conn = sqlite3.connect('fixed_asset.db')
        c = conn.cursor()
        c.execute('SELECT main_asset_id, asset_name, owner_name, location_2025, location_2026, "已盘点" FROM assets WHERE is_checked = 1')
        data = c.fetchall()
        conn.close()
        self.export_excel(data, "已盘点资产报表", "已盘点资产报表.xlsx")

    def export_unchecked(self):
        conn = sqlite3.connect('fixed_asset.db')
        c = conn.cursor()
        c.execute('SELECT main_asset_id, asset_name, owner_name, location_2025, location_2026, "未盘点" FROM assets WHERE is_checked = 0')
        data = c.fetchall()
        conn.close()
        self.export_excel(data, "未盘点资产报表", "未盘点资产报表.xlsx")

    def export_all(self):
        conn = sqlite3.connect('fixed_asset.db')
        c = conn.cursor()
        c.execute('SELECT main_asset_id, asset_name, owner_name, location_2025, location_2026, CASE WHEN is_checked = 1 THEN "已盘点" ELSE "未盘点" END FROM assets')
        data = c.fetchall()
        conn.close()
        self.export_excel(data, "全量资产盘点报表", "全量资产盘点报表.xlsx")

# -------------------------- 程序入口 --------------------------
def import_data_file(file_path):
    try:
        df = pd.read_excel(file_path, dtype={'主资产编号': str})
        required_cols = ['主资产编号', '资产名称', '姓名', '2025位置']
        if not all(col in df.columns for col in required_cols):
            print("错误：Excel文件列名不符合模板要求。需要包含：主资产编号、资产名称、姓名、2025位置")
            return False
        if '2026位置' not in df.columns:
            df['2026位置'] = ''
        df = df.drop_duplicates(subset=['主资产编号'], keep='first')

        # 将中文列名映射为数据库列名
        df = df.rename(columns={
            '主资产编号': 'main_asset_id',
            '资产名称': 'asset_name',
            '姓名': 'owner_name',
            '2025位置': 'location_2025',
            '2026位置': 'location_2026'
        })

        # 删除旧表（可能由错误的 to_sql 创建），并重建正确的表结构
        conn = sqlite3.connect('fixed_asset.db')
        c = conn.cursor()
        c.execute('DROP TABLE IF EXISTS assets')
        conn.commit()
        conn.close()
        init_db()
        conn = sqlite3.connect('fixed_asset.db')
        c = conn.cursor()

        rows = []
        for _, r in df.iterrows():
            main_id = str(r.get('main_asset_id') or '').strip()
            if not main_id:
                continue
            asset_name = r.get('asset_name') or ''
            owner_name = r.get('owner_name') or ''
            loc2025 = r.get('location_2025') or ''
            loc2026_raw = r.get('location_2026')
            # 处理 NaN/None
            if pd.isna(loc2026_raw):
                loc2026 = ''
            else:
                loc2026 = str(loc2026_raw).strip()
            is_checked = 1 if loc2026 != '' else 0
            rows.append((main_id, asset_name, owner_name, loc2025, loc2026, is_checked))

        c.executemany('INSERT OR REPLACE INTO assets (main_asset_id, asset_name, owner_name, location_2025, location_2026, is_checked) VALUES (?,?,?,?,?,?)', rows)
        conn.commit()
        conn.close()
        print(f"导入完成，共导入 {len(rows)} 条记录")
        return True
    except Exception as e:
        print(f"导入失败：{e}")
        return False


def export_excel_file(data, title, save_path):
    if len(data) == 0:
        print("暂无符合条件的资产数据")
        return False
    wb = Workbook()
    ws = wb.active
    ws.title = title
    headers = ['主资产编号', '资产名称', '姓名', '2025位置', '2026位置', '盘点状态']
    header_fill = PatternFill('solid', fgColor='0070C0')
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    center_align = Alignment(horizontal='center', vertical='center')
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    zebra_blue = 'EBF1F8'
    white = 'FFFFFF'
    for row_idx, row_data in enumerate(data, 2):
        bg = zebra_blue if (row_idx - 2) % 2 == 0 else white
        fill = PatternFill('solid', fgColor=bg)
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='left' if col_idx in [2,3,4,5] else 'center', vertical='center')
            cell.fill = fill
            if col_idx == 6:
                if value == '已盘点':
                    cell.font = Font(bold=True, color='00B050')
                else:
                    cell.font = Font(bold=True, color='FF0000')
    auto_fit_columns(ws)
    ws.freeze_panes = 'A2'
    wb.save(save_path)
    print(f"{title} 已导出到：{save_path}")
    return True


def export_checked_file(save_path):
    conn = sqlite3.connect('fixed_asset.db')
    c = conn.cursor()
    c.execute('SELECT main_asset_id, asset_name, owner_name, location_2025, location_2026, "已盘点" FROM assets WHERE is_checked = 1')
    data = c.fetchall()
    conn.close()
    return export_excel_file(data, "已盘点资产报表", save_path)


def export_unchecked_file(save_path):
    conn = sqlite3.connect('fixed_asset.db')
    c = conn.cursor()
    c.execute('SELECT main_asset_id, asset_name, owner_name, location_2025, location_2026, "未盘点" FROM assets WHERE is_checked = 0')
    data = c.fetchall()
    conn.close()
    return export_excel_file(data, "未盘点资产报表", save_path)


def export_all_file(save_path):
    conn = sqlite3.connect('fixed_asset.db')
    c = conn.cursor()
    c.execute('SELECT main_asset_id, asset_name, owner_name, location_2025, location_2026, CASE WHEN is_checked = 1 THEN "已盘点" ELSE "未盘点" END FROM assets')
    data = c.fetchall()
    conn.close()
    return export_excel_file(data, "全量资产盘点报表", save_path)


def headless_main():
    parser = argparse.ArgumentParser(description='固定资产盘点（无界面模式）')
    parser.add_argument('--init-db', action='store_true', help='初始化数据库')
    parser.add_argument('--import', dest='import_file', help='导入Excel文件')
    parser.add_argument('--export-checked', dest='export_checked', help='导出已盘点报表到指定路径')
    parser.add_argument('--export-unchecked', dest='export_unchecked', help='导出未盘点报表到指定路径')
    parser.add_argument('--export-all', dest='export_all', help='导出全量报表到指定路径')
    args = parser.parse_args()
    if args.init_db:
        init_db()
        print('数据库已初始化')
    if args.import_file:
        import_data_file(args.import_file)
    if args.export_checked:
        export_checked_file(args.export_checked)
    if args.export_unchecked:
        export_unchecked_file(args.export_unchecked)
    if args.export_all:
        export_all_file(args.export_all)


if __name__ == "__main__":
    init_db()
    if HAS_TK:
        root = tk.Tk()
        app = AssetCheckApp(root)
        root.mainloop()
    else:
        print('检测到无 GUI 环境，切换到无界面（headless）模式。使用 --help 查看可用命令。')
        headless_main()
